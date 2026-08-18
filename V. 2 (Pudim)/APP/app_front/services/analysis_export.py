"""Exportação auditável do FinScore Pudim 2.0.20."""

from __future__ import annotations

from datetime import datetime
import hashlib
from io import BytesIO
import json
import re
import unicodedata
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from finscore_v2 import core, executar_autotestes
except ModuleNotFoundError:
    from app_front.finscore_v2 import core, executar_autotestes


SHEET_ORDER = [
    "resumo_modelo", "confiabilidade", "correcoes_auditoria",
    "alertas_vies_material", "qualidade_dados", "rastreabilidade_contas",
    "contas_reportadas", "contas_utilizadas", "evidencia_serasa",
    "autotestes", "configuracao", "score_observado", "score_temporal",
    "contribuicoes_score", "caps_prudenciais", "intervalos_incerteza",
    "contas_derivadas", "indices_observados", "notas_observadas",
    "motivos_nan", "diagnostico_pca", "pesos_pca", "cargas_pca",
    "cenarios_deterministicos", "redundancia_fp", "comparacao_monte_carlo",
    "aceitos_rejeitados", "simulacoes_independentes",
    "simulacoes_correlacionadas", "resumo_simulacao", "simulacoes",
    "springate", "fleuriet_simplificado", "sensibilidade", "amplitudes",
]
_AUTOTESTES_CACHE: pd.DataFrame | None = None


def _autotestes() -> pd.DataFrame:
    global _AUTOTESTES_CACHE
    if _AUTOTESTES_CACHE is None:
        _AUTOTESTES_CACHE = executar_autotestes().copy(deep=True)
    return _AUTOTESTES_CACHE.copy(deep=True)


def _table(output: dict[str, Any], key: str) -> pd.DataFrame:
    value = output.get(key)
    if not isinstance(value, pd.DataFrame):
        return pd.DataFrame()
    result = value.copy(deep=True)
    result.attrs = {}
    return result


def _cargas_pca_orientadas(table: pd.DataFrame) -> pd.DataFrame:
    """Canonicaliza apenas o sinal arbitrário das cargas para o XLSX."""
    result = table.copy(deep=True)
    if "nucleo" not in result.columns:
        return result
    for nucleus in result["nucleo"].dropna().unique():
        mask = result["nucleo"].eq(nucleus)
        for component in [column for column in result if str(column).startswith("PC")]:
            values = pd.to_numeric(result.loc[mask, component], errors="coerce")
            if values.notna().any() and values.loc[values.abs().idxmax()] < 0:
                result.loc[mask, component] = -values
    return result


def _hash_regras() -> str:
    parameters = {
        "versao": core.VERSAO_MODELO,
        "hash_codigo": core.HASH_CODIGO_MODELO,
        "politica_correcao": core.POLITICA_CORRECAO,
        "limiar_confianca_automatica": core.LIMIAR_CONFIANCA_AUTOMATICA,
        "limiar_materialidade": core.LIMIAR_MATERIALIDADE,
        "limiar_vies_alto": core.LIMIAR_VIES_ALTO,
        "tolerancia_balanco": core.BALANCE_TOLERANCE,
        "cobertura_minima_nucleo": core.MIN_NUCLEUS_COVERAGE,
        "participacao_pca": core.PCA_ADAPTIVE_SHARE,
        "repeticoes_estabilidade_pca": core.PCA_STABILITY_REPETITIONS,
        "limiar_l1_pca": core.PCA_MAX_MEAN_L1_DISTANCE,
        "limiar_cosseno_pca": core.PCA_MIN_MEAN_COSINE,
        "pesos_nucleos": core.NUCLEUS_WEIGHTS,
        "peso_gargalo": core.BOTTLENECK_SHARE,
        "pesos_temporais": core.TEMPORAL_COMPONENT_WEIGHTS,
        "caps_prudenciais": core.PRUDENTIAL_CAPS,
        "ancoras": core.ANCHORS,
        "pesos_fixos": core.FIXED_WEIGHTS,
        "regra_excesso_fontes": core.EXCESS_SOURCE_RULE,
        "limite_rejeicao_mc": core.MAX_REJECTION_RATE,
        "persistencia_comum_mc": core.MC_TEMPORAL_PERSISTENCE,
        "cargas_comuns_mc": core.MC_COMMON_LOADINGS,
        "limiar_redundancia_fp": core.FP_REDUNDANCY_MATERIALITY_POINTS,
        "springate_ponto_corte": core.SPRINGATE_PONTO_CORTE,
        "fleuriet_escopo": core.FLEURIET_ESCOPO,
        "multiplicadores_prejuizo_recorrente": core.RECURRING_LOSS_MULTIPLIERS,
        "taxa_juros_efetiva_maxima": core.MAX_EFFECTIVE_INTEREST_RATE,
        "indicadores_por_nucleo": core.NUCLEI,
    }
    return hashlib.sha256(
        json.dumps(parameters, sort_keys=True).encode("utf-8")
    ).hexdigest()


def _resumo_modelo(output: dict[str, Any]) -> pd.DataFrame:
    status = output.get("status_qualidade", {})
    score = output.get("finscore_observado", {})
    reliability = output.get("confiabilidade", {})
    fields = [
        "Versão", "Status do cálculo", "Classificação de uso",
        "Apto para decisão", "Natureza do score", "Índice de confiabilidade",
        "Classificação da confiabilidade", "Correções/quarentenas aplicadas",
        "Pendências de confirmação", "Alertas de viés alto/crítico",
        "Score EO estrutural", "Score FP estrutural",
        "FinScore estrutural geométrico", "FinScore adaptativo geométrico",
        "FinScore prudencial antes do cap", "Cap prudencial aplicável",
        "FinScore prudencial final", "Faixa de incerteza inferior",
        "Faixa de incerteza superior", "Hash da base reportada",
        "Hash da base utilizada", "Hash das regras", "Hash do código",
    ]
    values = [
        output.get("modelo", {}).get("versao"), status.get("status"),
        status.get("classificacao_uso"), "SIM" if status.get("apto_decisao") else "NÃO",
        score.get("natureza_resultado", "NÃO CALCULADO"),
        status.get("indice_confiabilidade"), reliability.get("classificacao_confiabilidade"),
        status.get("correcoes_aplicadas", 0), status.get("correcoes_pendentes_confirmacao", 0),
        status.get("alertas_vies_alto_critico", 0), score.get("nucleo_EO_estrutural", np.nan),
        score.get("nucleo_FP_estrutural", np.nan), score.get("finscore_estrutural", np.nan),
        score.get("finscore_adaptativo", np.nan), score.get("finscore_prudencial_pre_cap", np.nan),
        score.get("cap_prudencial_aplicavel", np.nan), score.get("finscore_prudencial", np.nan),
        score.get("faixa_incerteza_inferior", np.nan), score.get("faixa_incerteza_superior", np.nan),
        output.get("hash_dados_reportados"), output.get("hash_dados_utilizados"),
        _hash_regras(), output.get("modelo", {}).get("hash_codigo"),
    ]
    return pd.DataFrame({"campo": fields, "valor": values})


def _configuracao(output: dict[str, Any], meta: dict[str, Any]) -> pd.DataFrame:
    model = output.get("modelo", {})
    status = output.get("status_qualidade", {})
    processed = model.get("processado_em")
    processed_text = processed.strftime("%Y-%m-%d %H:%M:%S") if isinstance(processed, datetime) else str(processed or "")
    rows = [
        ("versao", model.get("versao")), ("hash_codigo_modelo", model.get("hash_codigo")),
        ("planilha", meta.get("arquivo_origem", "upload Streamlit")), ("aba", "lancamentos"),
        ("simulacoes", model.get("numero_simulacoes", 0)), ("semente", model.get("semente")),
        ("data_hora_processamento", processed_text), ("politica_correcao", core.POLITICA_CORRECAO),
        ("aplicar_correcoes_automaticas", core.APLICAR_CORRECOES_AUTOMATICAS),
        ("limiar_confianca_automatica", core.LIMIAR_CONFIANCA_AUTOMATICA),
        ("limiar_materialidade", core.LIMIAR_MATERIALIDADE), ("limiar_vies_alto", core.LIMIAR_VIES_ALTO),
        ("delta_min", core.DELTA_MIN), ("delta_max", core.DELTA_MAX),
        ("regra_excesso_fontes", core.EXCESS_SOURCE_RULE), ("limite_rejeicao_mc", core.MAX_REJECTION_RATE),
        ("persistencia_comum_mc", core.MC_TEMPORAL_PERSISTENCE),
        ("persistencia_especifica_mc", core.MC_IDIOSYNCRATIC_PERSISTENCE),
        ("limiar_redundancia_fp_pontos", core.FP_REDUNDANCY_MATERIALITY_POINTS),
        ("springate_ponto_corte", core.SPRINGATE_PONTO_CORTE),
        ("fleuriet_escopo", core.FLEURIET_ESCOPO),
        ("multiplicadores_prejuizo_recorrente", str(core.RECURRING_LOSS_MULTIPLIERS)),
        ("taxa_juros_efetiva_maxima", core.MAX_EFFECTIVE_INTEREST_RATE),
        ("peso_EO", core.NUCLEUS_WEIGHTS["EO"]), ("peso_FP", core.NUCLEUS_WEIGHTS["FP"]),
        ("peso_gargalo", core.BOTTLENECK_SHARE), ("participacao_pca", core.PCA_ADAPTIVE_SHARE),
        ("pca_variaveis_minimas", core.PCA_MIN_ACTIVE_VARIABLES),
        ("pca_repeticoes_estabilidade", core.PCA_STABILITY_REPETITIONS),
        ("pca_l1_maximo", core.PCA_MAX_MEAN_L1_DISTANCE),
        ("pca_cosseno_minimo", core.PCA_MIN_MEAN_COSINE),
        ("peso_temporal_nivel", core.TEMPORAL_COMPONENT_WEIGHTS["nivel_atual"]),
        ("peso_temporal_dinamica", core.TEMPORAL_COMPONENT_WEIGHTS["dinamica_temporal"]),
        ("peso_temporal_resiliencia", core.TEMPORAL_COMPONENT_WEIGHTS["resiliencia"]),
        ("cobertura_minima_nucleo", core.MIN_NUCLEUS_COVERAGE),
        ("proxy_juros", core.USAR_DESPESAS_FINANCEIRAS_COMO_PROXY_JUROS),
        ("status_base", status.get("status")), ("classificacao_uso", status.get("classificacao_uso")),
        ("indice_confiabilidade", status.get("indice_confiabilidade")),
        ("hash_dados_reportados", output.get("hash_dados_reportados")),
        ("hash_dados_utilizados", output.get("hash_dados_utilizados")),
        ("hash_regras", _hash_regras()),
    ]
    return pd.DataFrame(rows, columns=["parametro", "valor"])


def montar_abas_exportacao(output: dict[str, Any], meta: dict[str, Any] | None = None) -> dict[str, pd.DataFrame]:
    """Monta as mesmas 35 abas e na mesma ordem do notebook 2.0.20."""
    meta = meta or {}
    sheets = {
        "resumo_modelo": _resumo_modelo(output),
        "confiabilidade": _table(output, "df_confiabilidade_componentes"),
        "correcoes_auditoria": _table(output, "df_correcoes_auditoria"),
        "alertas_vies_material": _table(output, "df_alertas_vies"),
        "qualidade_dados": _table(output, "df_qualidade"),
        "rastreabilidade_contas": _table(output, "df_rastreabilidade_contas"),
        "contas_reportadas": _table(output, "df_contas_reportadas"),
        "contas_utilizadas": _table(output, "df_contas_analise"),
        "evidencia_serasa": _table(output, "df_serasa"),
        "autotestes": _autotestes(),
        "configuracao": _configuracao(output, meta),
    }
    if output.get("status_qualidade", {}).get("apto_calculo", False):
        sheets.update({
            "score_observado": pd.DataFrame([output.get("finscore_observado", {})]),
            "score_temporal": _table(output, "df_score_temporal"),
            "contribuicoes_score": _table(output, "df_contribuicoes_score"),
            "caps_prudenciais": _table(output, "df_caps_prudenciais"),
            "intervalos_incerteza": _table(output, "df_intervalos_incerteza"),
            "contas_derivadas": _table(output, "df_contas_derivadas"),
            "indices_observados": _table(output, "df_indices_observados"),
            "notas_observadas": _table(output, "df_notas_observadas"),
            "motivos_nan": _table(output, "df_motivos_nan"),
            "diagnostico_pca": _table(output, "df_diagnostico_pca"),
            "pesos_pca": _table(output, "df_pesos_pca"),
            "cargas_pca": _cargas_pca_orientadas(_table(output, "df_cargas_pca")),
            "cenarios_deterministicos": _table(output, "df_cenarios_deterministicos"),
            "redundancia_fp": _table(output, "df_sensibilidade_redundancia_fp"),
            "comparacao_monte_carlo": _table(output, "df_comparacao_monte_carlo"),
            "aceitos_rejeitados": _table(output, "df_comparacao_aceitos_rejeitados"),
            "simulacoes_independentes": _table(output, "df_simulacoes_independentes"),
            "simulacoes_correlacionadas": _table(output, "df_simulacoes_correlacionadas"),
            "resumo_simulacao": _table(output, "df_resumo_simulacoes"),
            "simulacoes": _table(output, "df_simulacoes"),
            "springate": _table(output, "df_springate_complementar"),
            "fleuriet_simplificado": _table(output, "df_fleuriet_complementar"),
            "sensibilidade": _table(output, "df_sensibilidade"),
            "amplitudes": _table(output, "df_amplitudes"),
        })
    return sheets


def _style_workbook(workbook) -> None:
    navy, red, amber, green, blue = "17365D", "F4CCCC", "FCE5CD", "D9EAD3", "D9EAF7"
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[1].height = 34
        for column_cells in worksheet.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells[:200]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 11), 55)
        headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
        for label in ["severidade", "potencial_vies", "risco_vies", "status", "status_acao", "classificacao_uso"]:
            if label not in headers:
                continue
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, headers[label]); text = str(cell.value).upper()
                fill = red if any(x in text for x in ["CRIT", "BLOQUE", "NAO APTA"]) else amber if any(x in text for x in ["ALTO", "PROVIS", "PEND"]) else green if any(x in text for x in ["PASS", "DECISORIO", "CONTROLADO"]) else blue if any(x in text for x in ["QUARENTENA", "INFO"]) else None
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)
        for header, column in headers.items():
            text = str(header).lower()
            if any(token in text for token in ["valor_", "delta_absoluto", "score", "observado", "media", "mediana", "minimo", "maximo", "cap"]):
                for row in range(2, worksheet.max_row + 1): worksheet.cell(row, column).number_format = '#,##0.00;[Red](#,##0.00);-'
            if any(token in text for token in ["percentual", "materialidade", "confianca", "correlacao", "variancia", "participacao", "peso", "freq_", "impacto_absoluto", "cobertura", "amplitude", "choque"]):
                for row in range(2, worksheet.max_row + 1): worksheet.cell(row, column).number_format = "0.00%"
        if worksheet.title == "indices_observados":
            percentage = {"crescimento_receita", "margem_bruta", "margem_ebit", "margem_liquida", "capitalizacao", "endividamento_exigivel", "ccl_ativo", "ncg_operacional_ativo", "divida_liquida_ativo", "composicao_endividamento"}
            multiples = {"giro_ativo", "liquidez_corrente", "liquidez_seca", "cobertura_juros"}
            days = {"prazo_recebimento_dias", "prazo_estoques_dias", "prazo_fornecedores_dias", "ciclo_conversao_caixa"}
            for header, column in headers.items():
                fmt = "0.00%" if header in percentage else "0.00x" if header in multiples else "0.00" if header in days else None
                if fmt:
                    for row in range(2, worksheet.max_row + 1): worksheet.cell(row, column).number_format = fmt
        heights = {"correcoes_auditoria": 72, "alertas_vies_material": 72, "qualidade_dados": 48, "motivos_nan": 48, "caps_prudenciais": 48, "intervalos_incerteza": 48}
        if worksheet.title in heights:
            for row in range(2, worksheet.max_row + 1): worksheet.row_dimensions[row].height = heights[worksheet.title]
    if "correcoes_auditoria" in workbook.sheetnames:
        comments = {"valor_original": "Valor lido da planilha-fonte, nunca sobrescrito.", "valor_utilizado": "Valor empregado somente na cópia analítica.", "confianca": "Confiança da regra; não equivale a probabilidade estatística.", "materialidade_pct_ativo": "Magnitude da mudança dividida pelo Ativo Total do exercício.", "confirmado": "Verdadeiro somente após verificação documental."}
        for cell in workbook["correcoes_auditoria"][1]:
            if cell.value in comments: cell.comment = Comment(comments[cell.value], "FinScore")


def gerar_planilha_analise(output: dict[str, Any], meta: dict[str, Any] | None = None) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, table in montar_abas_exportacao(output, meta).items():
            table.to_excel(writer, sheet_name=name[:31], index=False)
        _style_workbook(writer.book)
    return buffer.getvalue()


def nome_arquivo_analise(output: dict[str, Any], meta: dict[str, Any] | None = None) -> str:
    company = str((meta or {}).get("empresa") or "empresa")
    normalized = unicodedata.normalize("NFKD", company).encode("ascii", "ignore").decode()
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", normalized).strip("_").lower() or "empresa"
    processed = output.get("modelo", {}).get("processado_em")
    timestamp = processed.strftime("%Y%m%d_%H%M") if isinstance(processed, datetime) else datetime.now().strftime("%Y%m%d_%H%M")
    version = str(output.get("modelo", {}).get("versao", "2.0.20"))
    return f"resultados_finscore_{version}_{slug}_{timestamp}.xlsx"
